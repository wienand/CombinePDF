import argparse
import datetime
import logging
import os
import shutil
import tempfile
from collections import defaultdict

import PyPDF2
import openpyxl
from PIL import Image


def check_file_type(filename):
    with open(filename, 'rb') as file:
        header = file.read(4)
        if header.startswith(b'%PDF'):
            return 'PDF'
        elif header[:3] == b'\xff\xd8\xff':
            return 'JPG'


def convert_jpg_2_pdf(source_file, pdf_file):
    img = Image.open(source_file)
    img = img.convert('RGB')
    img.save(pdf_file)


def combine_pdf_files(output_file_path, pdf_source_files):
    merger = PyPDF2.PdfMerger()
    for pdf_file in pdf_source_files:
        merger.append(pdf_file)
    merger.write(output_file_path)
    merger.close()


def combine_files(files_to_combine, root_path_input, root_path_output, **_):
    with tempfile.TemporaryDirectory() as temporary_directory:
        for output_file, source_files in files_to_combine.items():
            output_file_path = os.path.join(root_path_output, output_file)
            if not output_file_path.lower().endswith('.pdf'):
                output_file_path += '.pdf'
            if len(source_files) == 1:
                source_file = os.path.join(root_path_input, source_files[0])
                if check_file_type(source_file) == 'PDF':
                    logging.debug('Only one PDF, will just copy: %s --> %s', source_file, output_file_path)
                    # noinspection PyTypeChecker
                    shutil.copy(source_file, output_file_path)
                    continue
            pdf_source_files = []
            for source_file in source_files:
                source_file = os.path.join(root_path_input, source_file)
                file_type = check_file_type(source_file)
                if file_type == 'JPG':
                    logging.debug('JPG source, will convert to pdf: %s', source_file)
                    # noinspection PyTypeChecker
                    pdf_file = os.path.join(temporary_directory, source_file + '.PDF')
                    convert_jpg_2_pdf(source_file, pdf_file)
                    pdf_source_files.append(pdf_file)
                elif file_type == 'PDF':
                    pdf_source_files.append(source_file)
                else:
                    logging.warning('File is neither PDF nor JPG, skipping: %s', source_file)
            logging.debug('Combining PDF files to %s: %s', output_file_path, pdf_source_files)
            combine_pdf_files(output_file_path, pdf_source_files)


def read_excel(source_excel, source_sheet, column_input_files, column_output_file, **_):
    workbook = openpyxl.load_workbook(filename=source_excel)
    if source_sheet:
        if source_sheet in workbook.sheetnames:
            raise Exception(f'Sheet {source_sheet} does not exist in {source_excel}')
        sheet = workbook[source_sheet]
    else:
        sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    print(header)
    files_to_combine = defaultdict(list)
    for value in rows:
        row = dict(zip(header, value))
        if row[column_output_file] and row[column_input_files]:
            if row[column_input_files] not in files_to_combine[row[column_output_file]]:
                files_to_combine[row[column_output_file]].append(row[column_input_files])
    return files_to_combine


def main(args):
    # Read Excel file
    files_to_combine = read_excel(**args.__dict__)
    # Combine files
    combine_files(files_to_combine, **args.__dict__)


def parse_command_line():
    parser = argparse.ArgumentParser(description='Tool to combine pdf/jpf files to one pdf file based on ' +
                                                 'an excel list with resulting file names and source files')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('-v', '--verbose', action='store_true', help='be very verbose')
    group.add_argument('-q', '--quiet', action='store_true', help='no logging except errors')

    parser.add_argument('--source-excel', required=True,
                        help='Excel file containing ids and file locations')
    parser.add_argument('--source-sheet', required=False,
                        help='Sheet to read data from, if not given, active sheet is used')
    parser.add_argument('--column-input-files', required=True,
                        help='Column containing a path to the input files')
    parser.add_argument('--column-output-file', required=True,
                        help='Column containing a path to the output file the input files get combined to')

    parser.add_argument('--root-path-input', required=False, help='Root path for input file paths')
    parser.add_argument('--root-path-output', required=False, help='Root path for output file paths')

    return parser.parse_args()


if __name__ == '__main__':
    startTimestamp = datetime.datetime.now()
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(module)s %(levelname)s %(message)s')
    formatter = logging.Formatter('%(asctime)s %(module)s %(levelname)s %(message)s')
    fileHandler = logging.FileHandler('cpj.log', mode='a')
    fileHandler.setFormatter(formatter)
    logging.getLogger().addHandler(fileHandler)
    arguments = parse_command_line()
    if arguments.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    if arguments.quiet:
        logging.getLogger().setLevel(logging.ERROR)
    logging.debug('Command line parsed, begin processing ...')
    main(arguments)
